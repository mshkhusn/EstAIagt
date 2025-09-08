# app.py
import os
import re
import json
import importlib
from io import BytesIO
from datetime import date
import ast

import streamlit as st
import pandas as pd
import google.generativeai as genai
from dateutil.relativedelta import relativedelta

# ===== openpyxl / Excel =====
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter

# =========================
# ページ設定
# =========================
st.set_page_config(page_title="映像制作概算見積エージェント vNext", layout="centered")

# =========================
# Secrets
# =========================
GEMINI_API_KEY = st.secrets["GEMINI_API_KEY"]
OPENAI_API_KEY = st.secrets["OPENAI_API_KEY"]
APP_PASSWORD   = st.secrets["APP_PASSWORD"]

genai.configure(api_key=GEMINI_API_KEY)
os.environ["OPENAI_API_KEY"] = OPENAI_API_KEY

# =========================
# OpenAI 初期化（v1/v0 両対応）
# =========================
USE_OPENAI_CLIENT_V1 = False
openai_client = None
openai_version = "unknown"
try:
    from openai import OpenAI as _OpenAI
    openai_client = _OpenAI()
    USE_OPENAI_CLIENT_V1 = True
    try:
        mod = importlib.import_module("openai")
        openai_version = getattr(mod, "__version__", "1.x")
    except Exception:
        openai_version = "1.x"
except Exception:
    import openai as _openai
    _openai.api_key = OPENAI_API_KEY
    openai_client = _openai
    USE_OPENAI_CLIENT_V1 = False
    openai_version = getattr(openai_client, "__version__", "0.x")

# =========================
# 定数
# =========================
TAX_RATE = 0.10
MGMT_FEE_CAP_RATE = 0.15
RUSH_K = 0.75

# ===== 必須出現タクソノミー（網羅アンカー） =====
DETAIL_TAXONOMY = {
    "制作人件費": [
        "制作プロデューサー","制作プロジェクトマネージャー","制作進行","ディレクター",
        "アシスタントディレクター","カメラマン","撮影助手","照明","録音","DIT",
        "スタイリスト","ヘアメイク","美術デザイナー","大道具","小道具","ロケコーディネーター"
    ],
    "企画": [
        "企画構成","演出コンテ制作費","絵コンテ/コンテ","台本作成","ロケハン費","撮影許認可申請費"
    ],
    "撮影費": [
        "スタジオ費","ロケ費","車両/搬入出","カメラ機材","レンズ","照明機材","音声機材","ドローン","グリーンバック"
    ],
    "出演関連費": [
        "キャスト（メイン）","エキストラ","タレント使用料（該当時）","キャスティング費","衣装/フィッティング","交通/ケータリング"
    ],
    "編集費・MA費": [
        "オフライン編集","オンライン編集","カラコレ/グレーディング","VFX/CG","モーショングラフィックス",
        "字幕制作","MA","ナレーション収録/スタジオ","BGMライセンス/作曲"
    ],
    "諸経費": [
        "データ管理/バックアップ","予備日","保険","雑費/通信費","納品データ変換/複数書き出し"
    ],
    "管理費": ["管理費（固定）"]
}

# ===== スパース検出しきい値 =====
MIN_TOTAL_ITEMS = 20  # 管理費を除く最低行数
MIN_PER_CATEGORY = {
    "制作人件費": 6,
    "企画": 3,
    "撮影費": 5,
    "出演関連費": 3,
    "編集費・MA費": 5,
    "諸経費": 3
}

# ====== 相場レンジ（1日 or 1式単価の目安） ======
RATE_TABLE = {
    # 制作人件費
    "制作プロデューサー": {"unit":"日","low":70000,"mid":90000,"high":120000},
    "制作プロジェクトマネージャー": {"unit":"日","low":60000,"mid":80000,"high":100000},
    "制作進行": {"unit":"日","low":50000,"mid":65000,"high":80000},
    "ディレクター": {"unit":"日","low":90000,"mid":120000,"high":150000},
    "アシスタントディレクター": {"unit":"日","low":50000,"mid":60000,"high":75000},
    "カメラマン": {"unit":"日","low":70000,"mid":90000,"high":120000},
    "撮影助手": {"unit":"日","low":35000,"mid":45000,"high":60000},
    "照明": {"unit":"日","low":60000,"mid":80000,"high":100000},
    "録音": {"unit":"日","low":55000,"mid":70000,"high":90000},
    "DIT": {"unit":"日","low":45000,"mid":60000,"high":80000},
    "スタイリスト": {"unit":"日","low":55000,"mid":70000,"high":90000},
    "ヘアメイク": {"unit":"日","low":50000,"mid":65000,"high":80000},
    "美術デザイナー": {"unit":"日","low":65000,"mid":80000,"high":110000},
    "大道具": {"unit":"日","low":35000,"mid":45000,"high":60000},
    "小道具": {"unit":"日","low":30000,"mid":40000,"high":55000},
    "ロケコーディネーター": {"unit":"日","low":50000,"mid":65000,"high":85000},

    # 企画
    "企画構成": {"unit":"式","low":70000,"mid":90000,"high":130000},
    "演出コンテ制作費": {"unit":"式","low":50000,"mid":70000,"high":100000},
    "絵コンテ/コンテ": {"unit":"式","low":50000,"mid":70000,"high":100000},
    "台本作成": {"unit":"式","low":50000,"mid":70000,"high":100000},
    "ロケハン費": {"unit":"日","low":40000,"mid":50000,"high":70000},
    "撮影許認可申請費": {"unit":"式","low":20000,"mid":30000,"high":60000},

    # 撮影費
    "スタジオ費": {"unit":"日","low":60000,"mid":80000,"high":150000},
    "ロケ費": {"unit":"日","low":50000,"mid":70000,"high":120000},
    "車両/搬入出": {"unit":"日","low":30000,"mid":40000,"high":60000},
    "カメラ機材": {"unit":"日","low":50000,"mid":80000,"high":120000},
    "レンズ": {"unit":"日","low":20000,"mid":40000,"high":80000},
    "照明機材": {"unit":"日","low":40000,"mid":60000,"high":90000},
    "音声機材": {"unit":"日","low":20000,"mid":30000,"high":60000},
    "ドローン": {"unit":"日","low":80000,"mid":100000,"high":150000},
    "グリーンバック": {"unit":"日","low":30000,"mid":40000,"high":70000},

    # 出演関連費
    "キャスト（メイン）": {"unit":"人","low":30000,"mid":50000,"high":100000},
    "エキストラ": {"unit":"人","low":8000,"mid":12000,"high":20000},
    "タレント使用料（該当時）": {"unit":"式","low":200000,"mid":400000,"high":1000000},
    "キャスティング費": {"unit":"式","low":50000,"mid":70000,"high":120000},
    "衣装/フィッティング": {"unit":"式","low":20000,"mid":30000,"high":60000},
    "交通/ケータリング": {"unit":"日","low":15000,"mid":20000,"high":40000},

    # 編集・MA
    "オフライン編集": {"unit":"日","low":60000,"mid":80000,"high":120000},
    "オンライン編集": {"unit":"日","low":70000,"mid":90000,"high":140000},
    "カラコレ/グレーディング": {"unit":"日","low":60000,"mid":80000,"high":120000},
    "VFX/CG": {"unit":"日","low":80000,"mid":100000,"high":160000},
    "モーショングラフィックス": {"unit":"日","low":70000,"mid":90000,"high":140000},
    "字幕制作": {"unit":"式","low":20000,"mid":30000,"high":60000},
    "MA": {"unit":"日","low":50000,"mid":70000,"high":100000},
    "ナレーション収録/スタジオ": {"unit":"式","low":50000,"mid":70000,"high":120000},
    "BGMライセンス/作曲": {"unit":"式","low":40000,"mid":60000,"high":120000},

    # 諸経費
    "データ管理/バックアップ": {"unit":"式","low":15000,"mid":20000,"high":40000},
    "予備日": {"unit":"日","low":50000,"mid":60000,"high":80000},
    "保険": {"unit":"式","low":10000,"mid":15000,"high":30000},
    "雑費/通信費": {"unit":"式","low":8000,"mid":12000,"high":20000},
    "納品データ変換/複数書き出し": {"unit":"式","low":15000,"mid":20000,"high":40000},
}

CATEGORY_DEFAULT_UNIT = {
    "制作人件費": "日",
    "企画": "式",
    "撮影費": "日",
    "出演関連費": "人",
    "編集費・MA費": "日",
    "諸経費": "式",
    "管理費": "式",
}

# ====== 補正係数 ======
MARKET_MULT = {
    "日本全国平均": 1.00,
    "首都圏（東京近郊）": 1.15,
    "地方都市": 0.90,
}
CREW_GRADE_MULT = {
    "スタンダード": 1.00,
    "ジュニア": 0.90,
    "シニア": 1.20,
}
SCALE_MULT = {
    "小": 0.90,
    "中": 1.00,
    "大": 1.20,
}
FEATURE_MULT = {
    "CG": 1.10,
    "タレント": 1.15,
    "多数版書き出し": 1.05,
}

GLOBAL_FALLBACK_UNIT_PRICE = 30000  # 最終保険

# =========================
# セッション
# =========================
for k in ["items_json_raw", "items_json", "df", "meta", "final_html"]:
    if k not in st.session_state:
        st.session_state[k] = None

# =========================
# 認証
# =========================
st.title("映像制作概算見積エージェント vNext")
password = st.text_input("パスワードを入力してください", type="password")
if password != APP_PASSWORD:
    st.warning("🔒 認証が必要です")
    st.stop()

# =========================
# 入力
# =========================
st.header("制作条件の入力")
video_duration = st.selectbox("尺の長さ", ["15秒", "30秒", "60秒", "その他"])
final_duration = st.text_input("尺の長さ（自由記入）") if video_duration == "その他" else video_duration
num_versions = st.number_input("納品本数", min_value=1, max_value=10, value=1)
shoot_days = st.number_input("撮影日数", min_value=1, max_value=10, value=2)
edit_days = st.number_input("編集日数", min_value=1, max_value=10, value=3)
delivery_date = st.date_input("納品希望日", value=date.today() + relativedelta(months=1))
cast_main = st.number_input("メインキャスト人数", 0, 10, 1)
cast_extra = st.number_input("エキストラ人数", 0, 20, 0)
talent_use = st.checkbox("タレント起用あり")

default_roles = [
    "制作プロデューサー", "制作プロジェクトマネージャー", "ディレクター", "カメラマン",
    "照明スタッフ", "スタイリスト", "ヘアメイク"
]
selected_roles = st.multiselect("必要なスタッフ（選択式）", default_roles, default=default_roles)

custom_roles_text = st.text_input("その他のスタッフ（カンマ区切りで自由に追加）")
custom_roles = [role.strip() for role in custom_roles_text.split(",") if role.strip()]
staff_roles = selected_roles + custom_roles

shoot_location = st.text_input("撮影場所（例：都内スタジオ＋ロケ）")
kizai = st.multiselect("撮影機材", ["4Kカメラ", "照明", "ドローン", "グリーンバック"], default=["4Kカメラ", "照明"])
set_design_quality = st.selectbox("セット建て・美術装飾の規模", ["なし", "小（簡易装飾）", "中（通常レベル）", "大（本格セット）"])
use_cg = st.checkbox("CG・VFXあり")
use_narration = st.checkbox("ナレーション収録あり")
use_music = st.selectbox("音楽素材", ["既存ライセンス音源", "オリジナル制作", "未定"])
ma_needed = st.checkbox("MAあり")
deliverables = st.multiselect("納品形式", ["mp4（16:9）", "mp4（1:1）", "mp4（9:16）", "ProRes"])
subtitle_langs = st.multiselect("字幕言語", ["日本語", "英語", "その他"])
usage_region = st.selectbox("使用地域", ["日本国内", "グローバル", "未定"])
usage_period = st.selectbox("使用期間", ["3ヶ月", "6ヶ月", "1年", "2年", "無期限", "未定"])
budget_hint = st.text_input("参考予算（任意）")

# 備考 + 注意文
extra_notes = st.text_area("備考（案件概要・要件・想定媒体・必須/除外事項などを自由記入）")
st.caption("※備考に案件概要や条件を追記すると、不足項目の自動補完が働き、見積もりの精度が上がります。")

# 相場調整パラメータ
col1, col2, col3 = st.columns(3)
with col1:
    market = st.selectbox("相場地域", list(MARKET_MULT.keys()), index=0)
with col2:
    crew_grade = st.selectbox("人材グレード", list(CREW_GRADE_MULT.keys()), index=0)
with col3:
    scale = st.selectbox("制作規模", list(SCALE_MULT.keys()), index=1)

model_choice = st.selectbox("使用するAIモデル", ["Gemini 2.5 Pro", "GPT-5"])
do_normalize_pass = st.checkbox("LLMで正規化パスをかける（推奨）", value=True)
do_infer_from_notes = st.checkbox("備考から不足項目を推論して補完（推奨）", value=True)

# =========================
# ユーティリティ
# =========================
def join_or(value_list, empty="なし", sep=", "):
    if not value_list:
        return empty
    return sep.join(map(str, value_list))

def rush_coeff(base_days: int, target_days: int) -> float:
    if target_days >= base_days or base_days <= 0:
        return 1.0
    r = (base_days - target_days) / base_days
    return round(1 + RUSH_K * r, 2)

# ----- 数値ユーティリティ（カンマ・円マーク・全角対応） -----
def _to_float(v, default=0.0):
    if v is None:
        return float(default)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v)
    s = s.replace("￥", "").replace("¥", "").replace(",", "").replace("，", "").strip()
    z2h = str.maketrans("０１２３４５６７８９．－", "0123456789.-")
    s = s.translate(z2h)
    try:
        return float(s)
    except Exception:
        return float(default)

def _to_int(v, default=0):
    return int(round(_to_float(v, default)))

# ---------- JSON ロバストパース ----------
JSON_ITEMS_FALLBACK = {"items": []}

def _strip_code_fences(s: str) -> str:
    s = s.strip()
    if s.startswith("```"):
        s = re.sub(r"^```(json)?\s*", "", s, flags=re.IGNORECASE)
        s = re.sub(r"\s*```$", "", s)
    return s.strip()

def _remove_trailing_commas(s: str) -> str:
    return re.sub(r",\s*([}\]])", r"\1", s)

def _coerce_json_like(s: str):
    if not s:
        return None
    try:
        return json.loads(s)
    except Exception:
        pass
    try:
        first = s.find("{"); last = s.rfind("}")
        if first != -1 and last != -1 and last > first:
            frag = s[first:last+1]
            frag = _remove_trailing_commas(frag)
            frag2 = frag.replace("\r", "")
            frag2 = re.sub(r"\bTrue\b", "true", frag2)
            frag2 = re.sub(r"\bFalse\b", "false", frag2)
            frag2 = re.sub(r"\bNone\b", "null", frag2)
            if "'" in frag2 and '"' not in frag2:
                frag2 = frag2.replace("'", '"')
            try:
                return json.loads(frag2)
            except Exception:
                pass
    except Exception:
        pass
    try:
        return ast.literal_eval(s)
    except Exception:
        return None

def robust_parse_items_json(raw: str) -> str:
    s = _strip_code_fences(raw)
    obj = _coerce_json_like(s)
    if not isinstance(obj, dict):
        obj = JSON_ITEMS_FALLBACK.copy()
    items = obj.get("items")
    if not isinstance(items, list):
        if isinstance(obj.get("result"), dict) and isinstance(obj["result"].get("items"), list):
            items = obj["result"]["items"]
        elif isinstance(obj.get("data"), list):
            items = obj["data"]
        else:
            items = []
    obj["items"] = items
    return json.dumps(obj, ensure_ascii=False)

# ---------- 相場推定 ----------
def estimate_unit_price(task: str, category: str,
                        use="mid",
                        has_cg=False, has_talent=False,
                        deliverables_cnt=1,
                        market_key="日本全国平均",
                        grade_key="スタンダード",
                        scale_key="中") -> tuple[int, str]:
    base = RATE_TABLE.get(task)
    if not base:
        return (GLOBAL_FALLBACK_UNIT_PRICE, CATEGORY_DEFAULT_UNIT.get(category, "式"))
    unit = base["unit"]
    price = base.get(use, base["mid"])
    price *= MARKET_MULT.get(market_key, 1.0)
    price *= CREW_GRADE_MULT.get(grade_key, 1.0)
    price *= SCALE_MULT.get(scale_key, 1.0)
    if has_cg:
        price *= FEATURE_MULT["CG"]
    if has_talent:
        price *= FEATURE_MULT["タレント"]
    if deliverables_cnt and deliverables_cnt > 1:
        price *= FEATURE_MULT["多数版書き出し"]
    return (int(round(price, -2)), unit)

# ---------- プロンプト ----------
def _common_case_block() -> str:
    return f"""【案件条件】
- 尺: {final_duration}
- 本数: {num_versions}本
- 撮影日数: {shoot_days}日 / 編集日数: {edit_days}日
- 納品希望日: {delivery_date.isoformat()}
- キャスト: メイン{cast_main}人 / エキストラ{cast_extra}人 / タレント: {"あり" if talent_use else "なし"}
- スタッフ候補: {join_or(staff_roles, empty="未指定")}
- 撮影場所: {shoot_location if shoot_location else "未定"}
- 撮影機材: {join_or(kizai, empty="未指定")}
- 美術装飾: {set_design_quality}
- CG: {"あり" if use_cg else "なし"} / ナレーション: {"あり" if use_narration else "なし"} / 音楽: {use_music} / MA: {"あり" if ma_needed else "なし"}
- 納品形式: {join_or(deliverables, empty="未定")}
- 字幕: {join_or(subtitle_langs, empty="なし")}
- 使用地域: {usage_region} / 使用期間: {usage_period}
- 参考予算: {budget_hint if budget_hint else "未設定"}
- 備考: {extra_notes if extra_notes else "特になし"}"""

def _inference_block() -> str:
    if not do_infer_from_notes:
        return ""
    return """
- 備考や案件概要、一般的な広告映像制作の慣行から、未指定の必須/付随項目を**推論して必ず補完**すること。
  例: 企画構成、ロケハン、許認可申請、スタジオ/ロケ費、車両/機材搬入出、撮影助手、録音、DIT、メイキング、スチール、データ管理、CG/VFX、カラコレ、納品データ変換、権利処理、管理費など。
"""

def build_prompt_json() -> str:
    taxonomy_hint = "\n".join(
        f"- {cat}: " + ", ".join(DETAIL_TAXONOMY[cat]) for cat in DETAIL_TAXONOMY if cat != "管理費"
    )
    if model_choice == "GPT-5":
        return f"""
あなたは広告映像制作の見積り項目を作成するエキスパートです。
以下の条件を満たし、**JSONのみ**を返してください。

{_common_case_block()}

【出力仕様】
- JSON 1オブジェクト、ルートは items 配列のみ。
- 各要素キー: category / task / qty / unit / unit_price / note
- category は「制作人件費」「企画」「撮影費」「出演関連費」「編集費・MA費」「諸経費」「管理費」いずれか。
- **省略・統合を禁止**。似た名称でもまとめず、個別に列挙すること。
- **管理費を除いて最低 20 行以上**を出力。未知は妥当値で補完。
- 以下の**網羅アンカー**を最低限カバー（該当すれば各カテゴリから複数行を必ず含める）:
{taxonomy_hint}
{_inference_block()}
- 単価は一般的な日本の相場レンジ（low/mid/highの中央値相当）で推定し、0や未設定は入れないこと。
- qty は 0 禁止（>0）、unit_price は 0/負の値禁止。未確定でも妥当な推定値を入れること。
- qty/unit は現実的な単位（日/人/式/時間/カット等）、単価は日本の広告映像の一般レンジで推定。
- 管理費は固定1行（task=管理費（固定）, qty=1, unit=式）。
- 合計/税/HTMLなどは出力しない。
"""
    else:
        # Gemini も網羅アンカー＋統合禁止＋最低行数を適用
        return f"""
あなたは広告映像制作の見積り項目を作成するエキスパートです。
以下の条件を満たし、**JSONのみ**を返してください。

{_common_case_block()}

【出力仕様】
- JSON 1オブジェクト、ルートは items 配列のみ。
- 各要素キー: category / task / qty / unit / unit_price / note
- category は「制作人件費」「企画」「撮影費」「出演関連費」「編集費・MA費」「諸経費」「管理費」いずれか。
- **省略・統合の禁止**（似た名称でもまとめず個別行にする）。
- **管理費を除いて最低 20 行以上**を出力。未知は妥当値で補完。
- 以下の**網羅アンカー**を最低限カバー（該当すれば各カテゴリから複数行を必ず含める）:
{taxonomy_hint}
{_inference_block()}
- 単価は一般的な日本の相場レンジ（low/mid/highの中央値相当）で推定し、0や未設定は入れないこと。
- qty は 0 禁止（>0）、unit_price は 0/負の値禁止。未確定でも妥当な推定値を入れること。
- qty/unit は現実的な単位（日/人/式/時間/カット等）、単価は日本の広告映像の一般レンジで推定。
- 管理費は固定1行（task=管理費（固定）, qty=1, unit=式）。
- 合計/税/HTMLなどは出力しない。
"""

def build_normalize_prompt(items_json: str, preserve_detail: bool = False) -> str:
    if preserve_detail:
        return f"""
次のJSONを検査・正規化してください。返答は**修正済みJSONのみ**で、説明は不要です。
- スキーマ外キー削除、欠損補完（qty/unit/unit_price/note）。qty=0やunit_price<=0は禁止。妥当値で補完。
- **同義項目の統合や削減は禁止**（既存の粒度を保つ）
- category を次のいずれかへ正規化：制作人件費/企画/撮影費/出演関連費/編集費・MA費/諸経費/管理費
- 単位表記のゆれ（人日/日/式/本/時間/カット等）を正規化
- 管理費は固定1行（task=管理費（固定）, qty=1, unit=式）
【入力JSON】
{items_json}
"""
    return f"""
次のJSONを検査・正規化してください。返答は**修正済みJSONのみ**で、説明は不要です。
- スキーマ外キー削除、欠損補完
- category 正規化（制作人件費/企画/撮影費/出演関連費/編集費・MA費/諸経費/管理費）
- 単位正規化、同義項目統合、管理費は固定1行
【入力JSON】
{items_json}
"""

# ---------- LLM 呼び出し（JSON強制 & ロバストパース） ----------
def call_gpt_json(prompt: str) -> str:
    """GPT-5をJSONで強制返答。列挙性を高めるため若干のpresence_penaltyを付与。"""
    if USE_OPENAI_CLIENT_V1:
        resp = openai_client.chat.completions.create(
            model="gpt-5",
            messages=[
                {"role": "system", "content": "You MUST return a single valid JSON object only."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            temperature=0.7,
            presence_penalty=0.3,
        )
        return resp.choices[0].message.content
    else:
        resp = openai_client.ChatCompletion.create(
            model="gpt-5",
            messages=[
                {"role": "system", "content": "You MUST return a single valid JSON object only."},
                {"role": "user", "content": prompt}
            ],
            temperature=0.7,
            presence_penalty=0.3,
        )
        return resp["choices"][0]["message"]["content"]

def llm_generate_items_json(prompt: str) -> str:
    try:
        if model_choice == "Gemini 2.5 Pro":
            model = genai.GenerativeModel(
                "gemini-2.5-pro",
                generation_config={
                    "response_mime_type": "application/json",
                    "candidate_count": 1,
                    "max_output_tokens": 1600,
                }
            )
            res = model.generate_content(prompt).text
        else:
            res = call_gpt_json(prompt)
        st.session_state["items_json_raw"] = res
        return robust_parse_items_json(res)
    except Exception:
        # 最低限のフォールバック
        return json.dumps({"items":[
            {"category":"制作人件費","task":"制作プロデューサー","qty":1,"unit":"日","unit_price":80000,"note":"fallback"},
            {"category":"撮影費","task":"カメラマン","qty":max(1, int(shoot_days)),"unit":"日","unit_price":80000,"note":"fallback"},
            {"category":"編集費・MA費","task":"編集","qty":max(1, int(edit_days)),"unit":"日","unit_price":70000,"note":"fallback"},
            {"category":"管理費","task":"管理費（固定）","qty":1,"unit":"式","unit_price":120000,"note":"fallback"}
        ]}, ensure_ascii=False)

def llm_normalize_items_json(items_json: str) -> str:
    try:
        preserve = (model_choice == "GPT-5")
        prompt = build_normalize_prompt(items_json, preserve_detail=preserve)
        if model_choice == "Gemini 2.5 Pro":
            model = genai.GenerativeModel(
                "gemini-2.5-pro",
                generation_config={"response_mime_type": "application/json"}
            )
            res = model.generate_content(prompt).text
        else:
            res = call_gpt_json(prompt)
        return robust_parse_items_json(res)
    except Exception:
        return items_json

# ---------- “薄い出力なら追記して拡張”（再プロンプト） ----------
def _count_by_category(df: pd.DataFrame):
    if df.empty:
        return {}
    return df[df["category"] != "管理費"].groupby("category")["task"].count().to_dict()

def expand_if_sparse(items_json_str: str) -> str:
    """不足していたら“追記専用”で最大3回まで拡張。既存行は削除・上書きしない。"""
    def _needs_more(df: pd.DataFrame):
        if df.empty:
            return True, list(MIN_PER_CATEGORY.keys())
        non_mgmt = df[df["category"] != "管理費"]
        need_total = len(non_mgmt) < MIN_TOTAL_ITEMS
        counts = _count_by_category(df)
        need_cats = [cat for cat, mn in MIN_PER_CATEGORY.items() if counts.get(cat, 0) < mn]
        return (need_total or bool(need_cats)), need_cats

    try:
        df = df_from_items_json(items_json_str)
    except Exception:
        return items_json_str

    need, need_cats = _needs_more(df)
    if not need:
        return items_json_str

    taxonomy_hint = "\n".join(
        f"- {cat}: " + ", ".join(DETAIL_TAXONOMY[cat]) for cat in DETAIL_TAXONOMY if cat != "管理費"
    )

    attempts = 0
    while attempts < 3 and need:
        need_cats_str = ", ".join(need_cats) if need_cats else "全カテゴリ"
        prompt = f"""
次のJSON（現状の見積り）を**ベースに**、不足している項目を**追記のみ**して返してください。
- 既存項目は削除・統合・上書き禁止。**追記で拡張**すること。
- **管理費以外の合計行数が最低 {MIN_TOTAL_ITEMS} 行**になるまで追加。
- 不足カテゴリ：{need_cats_str}
- 参考タクソノミー（最低限カバー）:
{taxonomy_hint}
- 必要なら備考（案件概要）から合理的に推論して項目を補完する。
- 出力は**JSON 1オブジェクト（ルート items 配列のみ）**。

【現状JSON】
{items_json_str}
"""
        try:
            expanded = call_gpt_json(prompt)
            items_json_str = robust_parse_items_json(expanded)
            df = df_from_items_json(items_json_str)
            need, need_cats = _needs_more(df)
        except Exception:
            break
        attempts += 1

    return items_json_str

# ---------- df 生成（ゼロ禁止＋相場推定で補完） ----------
def df_from_items_json(items_json: str) -> pd.DataFrame:
    data = json.loads(items_json)
    items = data.get("items", [])
    rows = []

    has_cg = bool(use_cg)
    has_talent = bool(talent_use)
    deliverables_cnt = max(1, len(deliverables))

    for x in items:
        cat  = str(x.get("category","")).strip()
        task = str(x.get("task","")).strip()
        qty  = _to_float(x.get("qty", 0), 0)
        unit = str(x.get("unit","")).strip()
        price = _to_int(x.get("unit_price", 0), 0)
        note = str(x.get("note","")).strip()

        # 数量のデフォルト
        if qty <= 0:
            if cat == "出演関連費" and ("エキストラ" in task):
                qty = max(1, cast_extra or 1)
            elif cat == "出演関連費" and ("キャスト" in task):
                qty = max(1, cast_main or 1)
            else:
                qty = 1.0

        # 単価が無い/0なら相場から推定
        if price <= 0:
            price, unit_guess = estimate_unit_price(
                task, cat,
                use="mid",
                has_cg=has_cg, has_talent=has_talent,
                deliverables_cnt=deliverables_cnt,
                market_key=market,
                grade_key=crew_grade,
                scale_key=scale
            )
            if not unit:
                unit = unit_guess

        # 単位が空なら相場表/カテゴリ既定で補完
        if not unit:
            unit = RATE_TABLE.get(task, {}).get("unit") or CATEGORY_DEFAULT_UNIT.get(cat, "式")

        rows.append({
            "category": cat or "",
            "task": task or "",
            "qty": float(qty),
            "unit": unit,
            "unit_price": int(max(1, price)),
            "note": note,
        })

    return pd.DataFrame(rows)

# ---------- 計算 ----------
def compute_totals(df_items: pd.DataFrame, base_days: int, target_days: int):
    accel = rush_coeff(base_days, target_days)
    df_items = df_items.copy()
    df_items["小計"] = (df_items["qty"] * df_items["unit_price"]).round().astype(int)

    is_mgmt = (df_items["category"] == "管理費")
    df_items.loc[~is_mgmt, "小計"] = (df_items.loc[~is_mgmt, "小計"] * accel).round().astype(int)

    mgmt_current = int(df_items.loc[is_mgmt, "小計"].sum()) if is_mgmt.any() else 0
    subtotal_after_rush = int(df_items.loc[~is_mgmt, "小計"].sum())
    mgmt_cap = int(round(subtotal_after_rush * MGMT_FEE_CAP_RATE))
    mgmt_final = min(mgmt_current, mgmt_cap) if mgmt_current > 0 else mgmt_cap

    if is_mgmt.any():
        idx = df_items[is_mgmt].index[0]
        df_items.at[idx, "unit_price"] = mgmt_final
        df_items.at[idx, "qty"] = 1
        df_items.at[idx, "小計"] = mgmt_final
    else:
        df_items = pd.concat([df_items, pd.DataFrame([{
            "category":"管理費","task":"管理費（固定）","qty":1,"unit":"式","unit_price":mgmt_final,"小計":mgmt_final
        }])], ignore_index=True)

    taxable = int(df_items["小計"].sum())
    tax = int(round(taxable * TAX_RATE))
    total = taxable + tax

    meta = {
        "rush_coeff": accel,
        "subtotal_after_rush_excl_mgmt": subtotal_after_rush,
        "mgmt_fee_final": mgmt_final,
        "taxable": taxable,
        "tax": tax,
        "total": total,
    }
    return df_items, meta

def render_html(df_items: pd.DataFrame, meta: dict) -> str:
    def td_right(x): return f"<td style='text-align:right'>{x}</td>"
    html = []
    html.append("<p>以下は、映像制作にかかる各種費用をカテゴリごとに整理した概算見積書です。</p>")
    html.append(f"<p>短納期係数：{meta['rush_coeff']} ／ 管理費上限：{int(MGMT_FEE_CAP_RATE*100)}% ／ 消費税率：{int(TAX_RATE*100)}%</p>")
    html.append("<table border='1' cellspacing='0' cellpadding='6' style='border-collapse:collapse;width:100%'>")
    html.append("<thead><tr>"
                "<th style='text-align:left'>カテゴリ</th>"
                "<th style='text-align:left'>項目</th>"
                "<th style='text-align:right'>単価</th>"
                "<th style='text-align:left'>数量</th>"
                "<th style='text-align:left'>単位</th>"
                "<th style='text-align:right'>金額（円）</th>"
                "</tr></thead>")
    html.append("<tbody>")
    current_cat = None
    for _, r in df_items.iterrows():
        cat = r.get("category","")
        task = r.get("task","")
        unit_price_val = int(r.get("unit_price", 0) or 0)
        qty_val = r.get("qty", "")
        unit_val = r.get("unit","")
        subtotal_val = int(r.get("小計", 0) or 0)
        if cat != current_cat:
            html.append(f"<tr><td colspan='6' style='text-align:left;background:#f6f6f6;font-weight:bold'>{cat}</td></tr>")
            current_cat = cat
        html.append(
            "<tr>"
            f"<td>{cat}</td>"
            f"<td>{task}</td>"
            f"{td_right(f'{unit_price_val:,}')}"
            f"<td>{qty_val}</td>"
            f"<td>{unit_val}</td>"
            f"{td_right(f'{subtotal_val:,}')}"
            "</tr>"
        )
    html.append("</tbody></table>")
    html.append(
        f"<p><b>小計（税抜）</b>：{meta['taxable']:,}円　／　"
        f"<b>消費税</b>：{meta['tax']:,}円　／　"
        f"<b>合計</b>：<span style='color:red'>{meta['total']:,}円</span></p>"
    )
    html.append("<p>※本見積書は自動生成された概算です。実制作内容・条件により金額が増減します。</p>")
    return "\n".join(html)

def download_excel(df_items: pd.DataFrame, meta: dict):
    out = df_items.copy()
    out = out[["category","task","unit_price","qty","unit","小計"]]
    out.columns = ["カテゴリ","項目","単価（円）","数量","単位","金額（円）"]

    buf = BytesIO()
    try:
        import xlsxwriter  # noqa: F401
        engine = "xlsxwriter"
    except ModuleNotFoundError:
        engine = "openpyxl"

    with pd.ExcelWriter(buf, engine=engine) as writer:
        out.to_excel(writer, index=False, sheet_name="見積もり")

        if engine == "xlsxwriter":
            wb = writer.book
            ws = writer.sheets["見積もり"]
            fmt_int = wb.add_format({"num_format": "#,##0"})
            ws.set_column("A:B", 20)
            ws.set_column("C:C", 14, fmt_int)
            ws.set_column("D:D", 8)
            ws.set_column("E:E", 8)
            ws.set_column("F:F", 14, fmt_int)
            last_row = len(out) + 2
            ws.write(last_row,   4, "小計（税抜）")
            ws.write_number(last_row,   5, int(meta["taxable"]), fmt_int)
            ws.write(last_row+1, 4, "消費税")
            ws.write_number(last_row+1, 5, int(meta["tax"]), fmt_int)
            ws.write(last_row+2, 4, "合計")
            ws.write_number(last_row+2, 5, int(meta["total"]), fmt_int)
        else:
            ws = writer.book["見積もり"]
            widths = {"A":20, "B":20, "C":14, "D":8, "E":8, "F":14}
            for col, w in widths.items():
                ws.column_dimensions[col].width = w
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
                for cell in row: cell.number_format = '#,##0'
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
                for cell in row: cell.number_format = '#,##0'
            last_row = ws.max_row + 2
            ws.cell(row=last_row,   column=5, value="小計（税抜）")
            ws.cell(row=last_row,   column=6, value=int(meta["taxable"])).number_format = '#,##0'
            ws.cell(row=last_row+1, column=5, value="消費税")
            ws.cell(row=last_row+1, column=6, value=int(meta["tax"])).number_format = '#,##0'
            ws.cell(row=last_row+2, column=5, value="合計")
            ws.cell(row=last_row+2, column=6, value=int(meta["total"])).number_format = '#,##0'

    buf.seek(0)
    st.download_button("📥 Excelでダウンロード", buf, "見積もり.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# =========================
# DD見積書テンプレ出力（事前拡張テンプレ対応：行挿入なし）
# =========================
TOKEN_ITEMS = "{{ITEMS_START}}"
COLMAP = {
    "task": "B",        # 項目（B:N結合の左端セル）
    "qty": "O",         # 数量
    "unit": "Q",        # 単位
    "unit_price": "S",  # 単価
    "amount": "W",      # 金額（=O×S）結合の左上アンカー
}
BASE_START_ROW    = 19
BASE_SUBTOTAL_ROW = 72

def _find_token(ws, token: str):
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == token:
                return cell.row, cell.column
    return None, None

def _ensure_amount_formula(ws, row, qty_col_idx, price_col_idx, amount_col_idx):
    c = ws.cell(row=row, column=amount_col_idx)
    v = c.value
    if not (isinstance(v, str) and v.startswith("=")):
        qcol = get_column_letter(qty_col_idx)
        pcol = get_column_letter(price_col_idx)
        c.value = f"={qcol}{row}*{pcol}{row}"
    c.number_format = '#,##0'

def _update_subtotal_formula(ws, subtotal_row, start_row, end_row, amount_col_idx):
    ac = get_column_letter(amount_col_idx)
    if end_row < start_row:
        ws.cell(row=subtotal_row, column=amount_col_idx).value = 0
        ws.cell(row=subtotal_row, column=amount_col_idx).number_format = '#,##0'
    else:
        ws.cell(row=subtotal_row, column=amount_col_idx).value = f"=SUM({ac}{start_row}:{ac}{end_row})"
        ws.cell(row=subtotal_row, column=amount_col_idx).number_format = '#,##0'

def _find_subtotal_anchor_auto(ws, amount_col_idx: int):
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=amount_col_idx).value
        if isinstance(v, str) and v.startswith("=") and "SUM(" in v.upper():
            return r, amount_col_idx
    return None, None

def _write_preextended(ws, df_items: pd.DataFrame):
    r0, c0 = _find_token(ws, TOKEN_ITEMS)
    if r0:
        ws.cell(row=r0, column=c0).value = None
    start_row = r0 or BASE_START_ROW

    c_task = column_index_from_string(COLMAP["task"])
    c_qty  = column_index_from_string(COLMAP["qty"])
    c_unit = column_index_from_string(COLMAP["unit"])
    c_price= column_index_from_string(COLMAP["unit_price"])
    c_amt  = column_index_from_string(COLMAP["amount"])

    sub_r, _ = _find_subtotal_anchor_auto(ws, c_amt)
    if sub_r is None:
        sub_r = BASE_SUBTOTAL_ROW
    end_row = sub_r - 1
    capacity = end_row - start_row + 1
    n = len(df_items)

    if capacity <= 0:
        st.error("テンプレートの明細枠が不正です（小計行がITEMS_STARTより上にあります）。")
        return
    if n > capacity:
        st.warning(f"テンプレの明細枠（{capacity}行）を超えました。先頭から{capacity}行のみを書き込みます。")
        n = capacity

    for r in range(start_row, end_row + 1):
        cell_task = ws.cell(row=r, column=c_task)
        if not isinstance(cell_task, MergedCell):
            cell_task.value = None
        ws.cell(row=r, column=c_qty).value   = None
        ws.cell(row=r, column=c_unit).value  = None
        ws.cell(row=r, column=c_price).value = None
        _ensure_amount_formula(ws, r, c_qty, c_price, c_amt)

    for i in range(n):
        r = start_row + i
        row = df_items.iloc[i]
        ws.cell(row=r, column=c_task).value  = str(row.get("task",""))
        ws.cell(row=r, column=c_qty).value   = float(row.get("qty", 0) or 0)
        ws.cell(row=r, column=c_unit).value  = str(row.get("unit",""))
        ws.cell(row=r, column=c_price).value = int(float(row.get("unit_price", 0) or 0))

    last_detail_row = start_row + n - 1 if n > 0 else start_row - 1
    _update_subtotal_formula(ws, sub_r, start_row, last_detail_row, c_amt)

def export_with_template(template_bytes: bytes,
                         df_items: pd.DataFrame,
                         meta: dict):
    wb = load_workbook(filename=BytesIO(template_bytes))
    ws = wb.active
    _write_preextended(ws, df_items)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    st.download_button(
        "📥 DD見積書テンプレ（.xlsx）でダウンロード",
        out,
        "見積もり_DDテンプレ.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_dd_template"
    )

# =========================
# 実行
# =========================
if st.button("💡 見積もりを作成"):
    with st.spinner("AIが見積もり項目を作成中…"):
        prompt = build_prompt_json()
        items_json_str = llm_generate_items_json(prompt)

        if do_normalize_pass:
            items_json_str = llm_normalize_items_json(items_json_str)

        # ★ モデルに関わらず、不足なら追記専用で拡張（最大3回）
        items_json_str = expand_if_sparse(items_json_str)

        try:
            df_items = df_from_items_json(items_json_str)
        except Exception:
            st.error("JSONの解析に失敗しました。もう一度お試しください。")
            with st.expander("デバッグ：モデル生出力を見る"):
                st.code(st.session_state.get("items_json_raw", "(no raw)"))
            with st.expander("デバッグ：ロバスト整形後JSONを見る"):
                st.code(items_json_str, language="json")
            st.stop()

        base_days = int(shoot_days + edit_days + 5)
        target_days = (delivery_date - date.today()).days

        df_calc, meta = compute_totals(df_items, base_days, target_days)
        final_html = render_html(df_calc, meta)

        st.session_state["items_json"] = items_json_str
        st.session_state["df"] = df_calc
        st.session_state["meta"] = meta
        st.session_state["final_html"] = final_html

# =========================
# 表示 & ダウンロード
# =========================
if st.session_state["final_html"]:
    st.success("✅ 見積もり結果（サーバ計算で整合性チェック済み）")
    st.components.v1.html(st.session_state["final_html"], height=900, scrolling=True)
    download_excel(st.session_state["df"], st.session_state["meta"])

    st.markdown("---")
    st.subheader("DD見積書テンプレで出力")
    tmpl = st.file_uploader("DD見積書テンプレート（.xlsx）をアップロード", type=["xlsx"], key="tmpl_upload")
    if tmpl is not None:
        st.caption("テンプレに `{{ITEMS_START}}` を明細1行目（例：B19）に置いてください。小計セルはW列のSUM式で自動検出（例：W72）。行挿入は行いません。")
        export_with_template(
            tmpl.read(),
            st.session_state["df"],
            st.session_state["meta"]
        )

# =========================
# 開発者向け
# =========================
with st.expander("開発者向け情報（バージョン確認）", expanded=False):
    st.write({
        "openai_version": openai_version,
        "use_openai_client_v1": USE_OPENAI_CLIENT_V1,
        "infer_from_notes": do_infer_from_notes,
        "normalize_pass": do_normalize_pass,
        "model_choice": model_choice,
        "min_total_items": MIN_TOTAL_ITEMS,
        "min_per_category": MIN_PER_CATEGORY,
        "market": market,
        "crew_grade": crew_grade,
        "scale": scale,
    })
