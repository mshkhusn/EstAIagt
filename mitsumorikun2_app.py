# app.py （AI見積もりくん２ / GPT系のみ対応・広告制作全般カテゴリ例付き・DDテンプレ出力あり）

import os
import json
import importlib
from io import BytesIO
from datetime import date
import pandas as pd

import streamlit as st
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter

# ===== OpenAI v1 SDK =====
from openai import OpenAI
import httpx

# =========================
# ページ設定
# =========================
st.set_page_config(page_title="AI見積もりくん２", layout="centered")

# =========================
# Secrets
# =========================
OPENAI_API_KEY = st.secrets["OPENAI_API_KEY"]
APP_PASSWORD   = st.secrets["APP_PASSWORD"]
OPENAI_ORG_ID  = st.secrets.get("OPENAI_ORG_ID", None)

if not OPENAI_API_KEY:
    st.error("OPENAI_API_KEY が設定されていません。st.secrets を確認してください。")
    st.stop()

os.environ["OPENAI_API_KEY"] = OPENAI_API_KEY
if OPENAI_ORG_ID:
    os.environ["OPENAI_ORG_ID"] = OPENAI_ORG_ID

# OpenAI クライアント
openai_client = OpenAI(http_client=httpx.Client(timeout=60.0))

# =========================
# 定数
# =========================
TAX_RATE = 0.10

# =========================
# セッション管理
# =========================
for k in ["chat_history", "items_json_raw", "items_json", "df", "meta"]:
    if k not in st.session_state:
        st.session_state[k] = None

if st.session_state["chat_history"] is None:
    st.session_state["chat_history"] = [
        {"role": "system", "content": "あなたは広告クリエイティブ制作のプロフェッショナルです。相場感をもとに見積もりを作成するため、ユーザーにヒアリングを行います。"},
        {"role": "assistant", "content": "こんにちは！こちらは「AI見積もりくん２」です。見積もり作成のために、まず案件概要を教えてください。"}
    ]

# =========================
# 認証
# =========================
st.title("💰 AI見積もりくん２")
password = st.text_input("パスワードを入力してください", type="password")
if password != APP_PASSWORD:
    st.warning("🔒 認証が必要です")
    st.stop()

# =========================
# チャットUI
# =========================
st.header("チャットでヒアリング")

for msg in st.session_state["chat_history"]:
    if msg["role"] == "assistant":
        st.chat_message("assistant").write(msg["content"])
    elif msg["role"] == "user":
        st.chat_message("user").write(msg["content"])

if user_input := st.chat_input("要件を入力してください..."):
    st.session_state["chat_history"].append({"role": "user", "content": user_input})

    with st.chat_message("user"):
        st.write(user_input)

    with st.chat_message("assistant"):
        with st.spinner("AIが考えています..."):
            resp = openai_client.chat.completions.create(
                model="gpt-4.1",
                messages=st.session_state["chat_history"],
                temperature=0.4,
                max_tokens=1200
            )
            reply = resp.choices[0].message.content
            st.write(reply)
            st.session_state["chat_history"].append({"role": "assistant", "content": reply})

# =========================
# 見積もり生成用プロンプト
# =========================
def build_prompt_for_estimation(chat_history):
    return f"""
必ず有効な JSON のみを返してください。説明文は禁止です。

あなたは広告制作の見積もり作成エキスパートです。
以下の会話履歴をもとに、見積もりの内訳を作成してください。

【会話履歴】
{json.dumps(chat_history, ensure_ascii=False, indent=2)}

【カテゴリ例】
以下は広告制作でよく使われるカテゴリの例です：
- 企画・戦略関連（企画費、リサーチ費、コピーライティング、ディレクション など）
- デザイン・クリエイティブ制作（デザイン費、アートディレクション、イラスト制作 など）
- 撮影・映像関連（撮影費、スタッフ費、出演費、撮影機材費 など）
- 編集・仕上げ（編集費、CG/VFX、MA、字幕制作 など）
- Web関連（コーディング、CMS実装、テスト・QA、サーバー費 など）
- 配信・媒体関連（媒体出稿費、配信管理費、広告審査費 など）
- プロモーション・イベント関連（イベント運営費、会場費、施工費、スタッフ派遣 など）
- 諸経費・共通項目（交通費、宿泊費、消耗品費、雑費 など）
- 管理費（固定・一式）

【ルール】
- 上記カテゴリを参考にしつつ、案件内容に応じて適切に選択・追加・削除して構成してください。
- 各要素キー: category / task / qty / unit / unit_price / note
- qty, unit は妥当な値（日/式/人/時間/カット等）
- 単価は広告制作の一般相場で推定
- 「管理費」は必ず含める（task=管理費（固定）, qty=1, unit=式）
- 合計や税は含めない
"""

# =========================
# JSONパース
# =========================
def robust_parse_items_json(raw: str) -> str:
    try:
        obj = json.loads(raw)
        if not isinstance(obj, dict):
            obj = {"items": []}
        if "items" not in obj:
            obj["items"] = []
        return json.dumps(obj, ensure_ascii=False)
    except Exception:
        return json.dumps({"items":[]}, ensure_ascii=False)

# =========================
# 計算
# =========================
def df_from_items_json(items_json: str) -> pd.DataFrame:
    try:
        data = json.loads(items_json) if items_json else {}
    except Exception:
        data = {}
    items = data.get("items", []) or []
    norm = []
    for x in items:
        norm.append({
            "category": str(x.get("category", "")),
            "task": str(x.get("task", "")),
            "qty": x.get("qty", 0),
            "unit": str(x.get("unit", "")),
            "unit_price": x.get("unit_price", 0),
            "note": str(x.get("note", "")),
        })
    df = pd.DataFrame(norm)
    df["小計"] = (df["qty"].astype(float) * df["unit_price"].astype(float)).astype(int)
    return df

def compute_totals(df: pd.DataFrame):
    taxable = int(df["小計"].sum())
    tax = int(round(taxable * TAX_RATE))
    total = taxable + tax
    return {
        "taxable": taxable,
        "tax": tax,
        "total": total
    }

# =========================
# DDテンプレ出力
# =========================
TOKEN_ITEMS = "{{ITEMS_START}}"
COLMAP = {"task": "B", "qty": "O", "unit": "Q", "unit_price": "S", "amount": "W"}

def _find_token(ws, token: str):
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == token:
                return cell.row, cell.column
    return None, None

def _ensure_amount_formula(ws, row, qty_col_idx, price_col_idx, amount_col_idx):
    c = ws.cell(row=row, column=amount_col_idx)
    qcol = get_column_letter(qty_col_idx)
    pcol = get_column_letter(price_col_idx)
    c.value = f"={qcol}{row}*{pcol}{row}"
    c.number_format = '#,##0'

def _write_items_to_template(ws, df_items: pd.DataFrame):
    r0, c0 = _find_token(ws, TOKEN_ITEMS)
    if r0:
        ws.cell(row=r0, column=c0).value = None
    start_row = r0 or 19

    c_task = column_index_from_string(COLMAP["task"])
    c_qty  = column_index_from_string(COLMAP["qty"])
    c_unit = column_index_from_string(COLMAP["unit"])
    c_price= column_index_from_string(COLMAP["unit_price"])
    c_amt  = column_index_from_string(COLMAP["amount"])

    r = start_row
    current_cat = None

    for _, row in df_items.iterrows():
        cat = str(row.get("category", "")) or ""
        if cat != current_cat:
            ws.cell(row=r, column=c_task).value = cat
            ws.cell(row=r, column=c_task).font = Font(bold=True)
            _ensure_amount_formula(ws, r, c_qty, c_price, c_amt)
            current_cat = cat
            r += 1

        ws.cell(row=r, column=c_task).value  = str(row.get("task",""))
        ws.cell(row=r, column=c_qty).value   = float(row.get("qty", 0) or 0)
        ws.cell(row=r, column=c_unit).value  = str(row.get("unit",""))
        ws.cell(row=r, column=c_price).value = int(float(row.get("unit_price", 0) or 0))
        _ensure_amount_formula(ws, r, c_qty, c_price, c_amt)
        r += 1

def export_with_template(template_bytes: bytes, df_items: pd.DataFrame):
    wb = load_workbook(filename=BytesIO(template_bytes))
    ws = wb.active
    _write_items_to_template(ws, df_items)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# =========================
# 実行
# =========================
if st.button("📝 AI見積もりくんで見積もりを生成する"):
    with st.spinner("AIが見積もりを生成中…"):
        prompt = build_prompt_for_estimation(st.session_state["chat_history"])
        resp = openai_client.chat.completions.create(
            model="gpt-4.1",
            messages=[{"role":"system","content":"You MUST return only valid JSON."},
                      {"role":"user","content":prompt}],
            response_format={"type":"json_object"},
            temperature=0.2,
            max_tokens=4000
        )
        raw = resp.choices[0].message.content or '{"items":[]}'
        items_json = robust_parse_items_json(raw)
        df = df_from_items_json(items_json)
        meta = compute_totals(df)

        st.session_state["items_json_raw"] = raw
        st.session_state["items_json"] = items_json
        st.session_state["df"] = df
        st.session_state["meta"] = meta

# =========================
# 表示 & ダウンロード
# =========================
if st.session_state["df"] is not None:
    st.success("✅ 見積もり結果プレビュー")
    st.dataframe(st.session_state["df"])

    st.write(f"**小計（税抜）:** {st.session_state['meta']['taxable']:,}円")
    st.write(f"**消費税:** {st.session_state['meta']['tax']:,}円")
    st.write(f"**合計:** {st.session_state['meta']['total']:,}円")

    # Excel DL
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        st.session_state["df"].to_excel(writer, index=False, sheet_name="見積もり")
    buf.seek(0)
    st.download_button("📥 Excelでダウンロード", buf, "見積もり.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # DD見積書テンプレ
    tmpl = st.file_uploader("DD見積書テンプレートをアップロード（.xlsx）", type=["xlsx"])
    if tmpl is not None:
        out = export_with_template(tmpl.read(), st.session_state["df"])
        st.download_button("📥 DD見積書テンプレで出力", out, "見積もり_DDテンプレ.xlsx")
