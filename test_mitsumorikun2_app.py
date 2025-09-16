# app.py （AI見積もりくん２）
# GPT系のみ対応 / JSON強制 & 質問カテゴリフォールバック
# 追加要件込み再生成対応 / 追加質問時にプレビュー消去
# 見積もり生成後に「チャット入力欄の直上」にヒント文を必ず表示（st.emptyでプレースホルダ制御）

import os
import json
from io import BytesIO
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string, get_column_letter
from openai import OpenAI
import httpx

# --- 四隅インク（絶対パスで読んで、なければスキップ） ---
import base64
from pathlib import Path
import streamlit as st  # 既にimport済みなら重複OK

ROOT = Path(__file__).resolve().parent

def b64_or_none(p: Path) -> str:
    try:
        with p.open("rb") as f:
            return base64.b64encode(f.read()).decode()
    except Exception:
        # 無い／読めないときは空文字に
        return ""

INK_PINK   = b64_or_none(ROOT / "static" / "ink" / "ink_pink.png")
INK_CYAN   = b64_or_none(ROOT / "static" / "ink" / "ink_cyan.png")
INK_GREEN  = b64_or_none(ROOT / "static" / "ink" / "ink_green.png")
INK_PURPLE = b64_or_none(ROOT / "static" / "ink" / "ink_purple.png")

# 読めたものだけ重ねる（未使用なら削除可）
layers = []
if INK_PINK:   layers.append(f'url("data:image/png;base64,{INK_PINK}")   no-repeat left 3%  top 6%')
if INK_CYAN:   layers.append(f'url("data:image/png;base64,{INK_CYAN}")   no-repeat right 4% top 8%')
if INK_GREEN:  layers.append(f'url("data:image/png;base64,{INK_GREEN}")  no-repeat left 3%  bottom 6%')
if INK_PURPLE: layers.append(f'url("data:image/png;base64,{INK_PURPLE}") no-repeat right 4% bottom 5%')
bg_css = ",\n    ".join(layers) if layers else "none"
size_css = "220px 220px, 220px 220px, 220px 220px, 220px 220px" if layers else "auto"

# =========================
# ページ設定
# =========================
st.set_page_config(page_title="AI見積もりくん２", layout="centered")

# =========================
# デザイン一式
# =========================
st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Mochiy+Pop+One&display=swap');

/* ===== Base ===== */
html, body {{ background:#000 !important; }}
.stApp, .stApp * {{
  background:transparent !important;
  color:#fff !important;
  font-family:'Mochiy Pop One',sans-serif !important;
  letter-spacing:.01em;
}}

/* ヘッダー/サイドバー */
[data-testid="stHeader"],[data-testid="stToolbar"],[data-testid="stStatusWidget"],
[data-testid="stSidebar"],[data-testid="stSidebarContent"] {{
  background:transparent !important; border:none !important;
}}

/* ===== Inputs ===== */
.stTextInput label, .stTextArea label, .stSelectbox label {{ color:#fff important; }}
.stTextInput input, .stTextArea textarea, .stSelectbox div {{
  background:#111 !important; color:#fff !important;
  border:1px solid #555 !important; border-radius:10px !important;
}}
.stTextInput input::placeholder, .stTextArea textarea::placeholder,
.stChatInput textarea::placeholder {{ color:#ddd !important; }}

/* 目アイコン */
.stTextInput [data-baseweb="button"] {{
  background:#333 !important; color:#fff !important;
  border:1px solid #666 !important; border-radius:10px !important;
}}

/* ===== Buttons（デフォルト） ===== */
.stButton button, .stDownloadButton > button {{
  background:#222 !important; color:#fff !important;
  border:1px solid #666 !important; border-radius:10px !important;
  padding:.55rem 1rem !important; box-shadow:none !important;
}}
.stButton button:hover, .stDownloadButton > button:hover {{
  background:#2c2c2c !important; border-color:#777 !important;
}}

/* ===== Chat ===== */
[data-testid="stChatMessage"] {{
  background:transparent !important; border:none !important; border-radius:14px !important;
}}
[data-testid="stChatInput"], [data-testid="stChatInput"]>div {{ background:transparent !important; }}
.stChatInput textarea {{
  background:#111 !important; color:#fff !important;
  border:1px solid #555 !important; border-radius:10px !important;
}}
.stChatInput [data-baseweb="button"] {{
  background:#222 !important; color:#fff !important;
  border:1px solid #555 !important; border-radius:10px !important;
}}

/* ===== File Uploader ===== */
[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] {{
  position:relative !important; background:#111 !important; color:#fff !important;
  border:1.5px solid #666 !important; border-radius:12px !important; padding-left:64px !important;
}}
[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] svg {{ display:none !important; }}
@supports selector(div:has(> svg)) {{
  [data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"] div:has(> svg) {{
    background:transparent !important; border:none !important;
  }}
}}
[data-testid="stFileUploader"] [data-testid="stFileUploaderDropzone"]::before {{
  content:""; position:absolute; left:18px; top:50%; transform:translateY(-50%);
  width:32px; height:32px; background-repeat:no-repeat; background-position:center; background-size:contain;
  background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' fill='%23ffffff' viewBox='0 0 24 24'%3E%3Cpath d='M6 19a4 4 0 0 1 0-8 5 5 0 0 1 9.7-1.4A3.5 3.5 0 1 1 18 19H6z'/%3E%3C/svg%3E");
}}

/* ===== Avatar ===== */
.stApp [data-testid="stChatMessage"] [data-testid*="Avatar"] {{
  background:#a64dff !important; color:#fff !important; border-radius:12px !important;
}}
.stApp [data-testid="stChatMessage"]:has([data-testid*="user"]) [data-testid*="Avatar"] {{
  background:#00e08a !important; color:#000 !important;
}}

/* ===== 見積もり結果見出し ===== */
.preview-title {{
  font-size:32px !important; font-weight:900 !important;
  color:#78f416 !important; margin-bottom:16px !important;
}}

/* ===== フォーカス演出 ===== */
.stChatInput:focus-within textarea, .stTextInput input:focus {{
  border:3px solid transparent !important; border-radius:12px !important; background:#111 !important;
  border-image:linear-gradient(90deg,#ff4df5,#90fb0f,#00c3ff) 1 !important;
  box-shadow:0 0 12px rgba(255,77,245,.6), 0 0 18px rgba(144,251,15,.5), 0 0 24px rgba(0,195,255,.4) !important;
}}

/* ===== Markdown テーブル・水平線 ===== */
[data-testid="stMarkdownContainer"] table {{ border-collapse:collapse !important; border:1px solid #fff !important; }}
[data-testid="stMarkdownContainer"] th, [data-testid="stMarkdownContainer"] td {{
  border:1px solid #fff !important; padding:6px 10px !important; color:#fff !important;
}}
[data-testid="stMarkdownContainer"] th {{ background-color:rgba(255,255,255,.1) !important; font-weight:700 !important; }}
[data-testid="stMarkdownContainer"] hr {{ border:none !important; border-top:1px solid #fff !important; margin:1em 0 !important; }}

/* ===== 旧 .stApp::before を無効化（二重防止） ===== */
.stApp::before {{ content:""; background:none !important; }}

/* ===== 四隅インク：ピンクさらに上／ブルー右にはみ出し／イエロー縦長 ===== */
body::before {{
  content:"";
  position: fixed;
  inset:0;
  background:
    url("data:image/png;base64,{INK_PINK}")   no-repeat left  -160px top  -160px, /* ピンク */
    url("data:image/png;base64,{INK_CYAN}")   no-repeat right -220px top  -60px,  /* ブルー → 右へ100px */
    url("data:image/png;base64,{INK_GREEN}")  no-repeat left  -100px bottom -100px;/* イエロー */
  background-size:
    380px 380px,   /* ピンク */
    500px 500px,   /* ブルー */
    260px 400px;   /* イエロー縦長 */
  pointer-events: none;
  z-index: -1;
}}
@media (max-width: 900px) {{
  body::before {{
    background-position:
      left  -80px top   -80px,   /* ピンク */
      right -160px top  -30px,   /* ブルー → モバイルも少し右へ */
      left -100px bottom -60px;  /* イエロー */
    background-size:
      260px 260px,   /* ピンク縮小 */
      320px 320px,   /* ブルー縮小 */
      200px 320px;   /* イエロー縦長 */
  }}
}}

/* ===== 生成ボタン：グリーン→ブルーのグラデ（コンテナ版） ===== */
/* 同じ縦ブロック（stVerticalBlock）内のどこかに .gen-scope があれば、そのブロックのボタンを着色 */
[data-testid="stVerticalBlock"]:has(.gen-scope) div.stButton > button {
  background: linear-gradient(90deg, #00e08a, #00c3ff) !important;
  color: #fff !important;
  border: none !important;
  border-radius: 12px !important;
  padding: .68rem 1.15rem !important;
  font-weight: 700 !important;
  text-shadow: 0 1px 0 rgba(0,0,0,.25);
  box-shadow:
    0 0 10px rgba(0,224,138,.55),
    0 0 18px rgba(0,195,255,.45) !important;
  transition: transform .08s ease, filter .15s ease, box-shadow .15s ease;
}
[data-testid="stVerticalBlock"]:has(.gen-scope) div.stButton > button:hover {
  filter: brightness(1.08);
  transform: translateY(-1px);
  box-shadow:
    0 0 12px rgba(0,224,138,.65),
    0 0 24px rgba(0,195,255,.55) !important;
}
[data-testid="stVerticalBlock"]:has(.gen-scope) div.stButton > button:active {
  filter: brightness(.98);
  transform: translateY(0);
}
</style>
""", unsafe_allow_html=True)

# =========================
# Secrets
# =========================
OPENAI_API_KEY = st.secrets["OPENAI_API_KEY"]
APP_PASSWORD   = st.secrets["APP_PASSWORD"]
OPENAI_ORG_ID  = st.secrets.get("OPENAI_ORG_ID", None)

if not OPENAI_API_KEY:
    st.error("OPENAI_API_KEY が設定されていません。st.secrets を確認してください。")
    st.stop()

os.environ["OPENAI_API_KEY"] = OPENAI_API_KEY
if OPENAI_ORG_ID:
    os.environ["OPENAI_ORG_ID"] = OPENAI_ORG_ID

openai_client = OpenAI(http_client=httpx.Client(timeout=60.0))

# =========================
# 定数
# =========================
TAX_RATE = 0.10

# =========================
# セッション管理
# =========================
for k in ["chat_history", "items_json_raw", "items_json", "df", "meta"]:
    if k not in st.session_state:
        st.session_state[k] = None

if st.session_state["chat_history"] is None:
    st.session_state["chat_history"] = [
        {"role": "system", "content": "あなたは広告クリエイティブ制作のプロフェッショナルです。相場感をもとに見積もりを作成するため、ユーザーにヒアリングを行います。"},
        {"role": "assistant", "content": "こんにちは！こちらは「AI見積もりくん２」です。見積もり作成のために、まず案件概要を教えてください。"}
    ]

# =========================
# 認証（ロゴ）
# =========================
st.markdown("""
<style>
/* ===== ロゴ（潰れ防止：正規ウェイト＋字間少し広め） ===== */
.logo-wrap{
  display:flex; justify-content:center; align-items:center;
  width:100%; margin:24px 0 40px 0;
}
.logo-pill{
  display:inline-block; padding:6px; border-radius:9999px !important;
  background:linear-gradient(90deg,#ff4df5,#a64dff) !important;
}
.logo-box{
  padding:30px 76px; border-radius:9999px !important; background:#000 !important;
  font-family:'Mochiy Pop One',sans-serif; color:inherit !important; white-space:nowrap;
  -webkit-font-smoothing:antialiased; text-rendering:optimizeLegibility;
}
.logo-row1{
  display:flex; align-items:flex-start; justify-content:center;
  gap:8px; line-height:1.02; margin:0;
}
.logo-box .ai{
  font-size:90px; font-weight:400; letter-spacing:0.5px; color:#ff4df5 !important;
}
.logo-box .mitsumori{
  font-size:60px; font-weight:400; letter-spacing:0.5px; color:#fff !important;
}
.logo-kunrow{
  text-align:center; line-height:1.0; margin-top:-22px; letter-spacing:0.5px;
}
.logo-box .kun{  font-size:44px; font-weight:400; color:#fff !important; }
.logo-box .num2{ font-size:44px; font-weight:400; color:#ff4df5 !important; }
</style>

<div class="logo-wrap">
  <div class="logo-pill">
    <div class="logo-box">
      <div class="logo-row1">
        <span class="ai">AI</span><span class="mitsumori">見積もり</span>
      </div>
      <div class="logo-kunrow">
        <span class="kun">くん</span><span class="num2">2</span>
      </div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

password = st.text_input("パスワードを入力してください", type="password")
if password != APP_PASSWORD:
    st.warning("認証が必要です")
    st.stop()

# =========================
# チャットUI（Markdownで描画）
# =========================
st.markdown("""
<style>
.custom-header {
  color: #90fb0f !important;
  font-size: 40px !important;
  font-weight: 400 !important;  /* Mochiy Pop One は 400 だけ */
  font-family: 'Mochiy Pop One', sans-serif !important;
  letter-spacing: 1px !important;   /* 潰れ防止に少し広げる */
  line-height: 1.3 !important;
  text-shadow: 0 0 4px rgba(0,0,0,0.6);
}
</style>
""", unsafe_allow_html=True)

st.markdown(
    '<h2 class="custom-header">AI見積もりくんにチャットで相談して見積もりを生成しよう！</h2>',
    unsafe_allow_html=True
)

# 既存履歴をMarkdownで再描画
for msg in st.session_state["chat_history"]:
    if msg["role"] == "assistant":
        st.chat_message("assistant").markdown(msg["content"])
    elif msg["role"] == "user":
        st.chat_message("user").markdown(msg["content"])

# --- ヒント文プレースホルダ（チャット入力直前） ---
hint_placeholder = st.empty()
if st.session_state["df"] is not None:
    hint_placeholder.caption(
        "チャットをさらに続けて見積もり精度を上げることができます。\n"
        "追加で要件を入力した後に再度このボタンを押すと、過去のチャット履歴＋新しい要件を反映して見積もりが更新されます。"
    )

# =========================
# 入力欄
# =========================
if user_input := st.chat_input("要件を入力してください..."):
    # 新しい入力があれば過去の見積もり結果をクリア（プレビューを一度消す）
    st.session_state["df"] = None
    st.session_state["meta"] = None
    st.session_state["items_json"] = None
    st.session_state["items_json_raw"] = None

    st.session_state["chat_history"].append({"role": "user", "content": user_input})

    with st.chat_message("user"):
        st.markdown(user_input)

    with st.chat_message("assistant"):
        with st.spinner("AIが考えています..."):
            resp = openai_client.chat.completions.create(
                model="gpt-4.1",
                messages=st.session_state["chat_history"],
                temperature=0.4,
                max_tokens=1200
            )
            reply = resp.choices[0].message.content
            st.markdown(reply)  # ← Markdownそのまま表示
            st.session_state["chat_history"].append({"role": "assistant", "content": reply})

# =========================
# 見積もり生成用プロンプト
# =========================
def build_prompt_for_estimation(chat_history):
    return f"""
必ず有効な JSON のみを返してください。説明文・文章・Markdown・テーブルは禁止です。

あなたは広告制作の見積もり作成エキスパートです。
以下の会話履歴をもとに、見積もりの内訳を作成してください。

【会話履歴】
{json.dumps(chat_history, ensure_ascii=False, indent=2)}

【カテゴリ例】
- 企画・戦略関連（企画費、リサーチ費、コピーライティング、ディレクション など）
- デザイン・クリエイティブ制作（デザイン費、アートディレクション、イラスト制作 など）
- 撮影・映像関連（撮影費、スタッフ費、出演費、撮影機材費 など）
- 編集・仕上げ（編集費、CG/VFX、MA、字幕制作 など）
- Web関連（コーディング、CMS実装、テスト・QA、サーバー費 など）
- 配信・媒体関連（媒体出稿費、配信管理費、広告審査費 など）
- プロモーション・イベント関連（イベント運営費、会場費、施工費、スタッフ派遣 など）
- 諸経費・共通項目（交通費、宿泊費、消耗品費、雑費 など）
- 管理費（固定・一式）

【ルール】
- 必ず items 配列には1行以上の見積もり項目を返してください（空配列は禁止）。
- 各要素キー: category / task / qty / unit / unit_price / note
- 欠損がある場合は補完してください。
- 「管理費」は必ず含める（task=管理費（固定）, qty=1, unit=式）。
- 合計や税は含めない。
- もし情報不足で正しい見積もりが作れない場合は、items に1行だけ
  {{"category":"質問","task":"追加で必要な情報を教えてください","qty":0,"unit":"","unit_price":0,"note":"不足情報あり"}}
  を返してください。
"""

# =========================
# JSONパース & フォールバック
# =========================
def robust_parse_items_json(raw: str) -> str:
    try:
        obj = json.loads(raw)
    except Exception:
        return json.dumps({
            "items":[
                {"category":"質問","task":"要件を詳しく教えてください","qty":0,"unit":"","unit_price":0,"note":"AIがテキストを返しました"}
            ]
        }, ensure_ascii=False)

    if not isinstance(obj, dict):
        obj = {"items":[]}
    if "items" not in obj or not obj["items"]:
        obj["items"] = [{
            "category":"質問","task":"追加で要件を教えてください","qty":0,"unit":"","unit_price":0,"note":"不足情報あり"
        }]
    return json.dumps(obj, ensure_ascii=False)

# =========================
# DataFrame生成
# =========================
def df_from_items_json(items_json: str) -> pd.DataFrame:
    try:
        data = json.loads(items_json) if items_json else {}
    except Exception:
        data = {}
    items = data.get("items", []) or []
    norm = []
    for x in items:
        norm.append({
            "category": str((x or {}).get("category", "")),
            "task": str((x or {}).get("task", "")),
            "qty": (x or {}).get("qty", 0) or 0,
            "unit": str((x or {}).get("unit", "")),
            "unit_price": (x or {}).get("unit_price", 0) or 0,
            "note": str((x or {}).get("note", "")),
        })
    df = pd.DataFrame(norm)
    for col in ["category", "task", "qty", "unit", "unit_price", "note"]:
        if col not in df.columns:
            df[col] = "" if col in ["category","task","unit","note"] else 0
    df["qty"] = pd.to_numeric(df["qty"], errors="coerce").fillna(0).astype(float)
    df["unit_price"] = pd.to_numeric(df["unit_price"], errors="coerce").fillna(0).astype(int)
    df["小計"] = (df["qty"] * df["unit_price"]).astype(int)
    return df

# =========================
# 合計計算
# =========================
def compute_totals(df: pd.DataFrame):
    taxable = int(df["小計"].sum())
    tax = int(round(taxable * TAX_RATE))
    total = taxable + tax
    return {"taxable": taxable, "tax": tax, "total": total}

# =========================
# DDテンプレ出力
# =========================
TOKEN_ITEMS = "{{ITEMS_START}}"
COLMAP = {"task": "B", "qty": "O", "unit": "Q", "unit_price": "S", "amount": "W"}

def _find_token(ws, token: str):
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == token:
                return cell.row, cell.column
    return None, None

def _ensure_amount_formula(ws, row, qty_col_idx, price_col_idx, amount_col_idx):
    c = ws.cell(row=row, column=amount_col_idx)
    qcol = get_column_letter(qty_col_idx)
    pcol = get_column_letter(price_col_idx)
    c.value = f"={qcol}{row}*{pcol}{row}"
    c.number_format = '#,##0'

def _write_items_to_template(ws, df_items: pd.DataFrame):
    r0, c0 = _find_token(ws, TOKEN_ITEMS)
    if r0:
        ws.cell(row=r0, column=c0).value = None
    start_row = r0 or 19

    c_task = column_index_from_string(COLMAP["task"])
    c_qty  = column_index_from_string(COLMAP["qty"])
    c_unit = column_index_from_string(COLMAP["unit"])
    c_price= column_index_from_string(COLMAP["unit_price"])
    c_amt  = column_index_from_string(COLMAP["amount"])

    r = start_row
    current_cat = None
    for _, row in df_items.iterrows():
        cat = str(row.get("category", "")) or ""
        if cat != current_cat:
            ws.cell(row=r, column=c_task).value = cat
            ws.cell(row=r, column=c_task).font = Font(bold=True)
            _ensure_amount_formula(ws, r, c_qty, c_price, c_amt)
            current_cat = cat
            r += 1
        ws.cell(row=r, column=c_task).value  = str(row.get("task",""))
        ws.cell(row=r, column=c_qty).value   = float(row.get("qty", 0) or 0)
        ws.cell(row=r, column=c_unit).value  = str(row.get("unit",""))
        ws.cell(row=r, column=c_price).value = int(float(row.get("unit_price", 0) or 0))
        _ensure_amount_formula(ws, r, c_qty, c_price, c_amt)
        r += 1

def export_with_template(template_bytes: bytes, df_items: pd.DataFrame):
    wb = load_workbook(filename=BytesIO(template_bytes))
    ws = wb.active
    _write_items_to_template(ws, df_items)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# =========================
# 実行（★ コンテナ方式でボタンにグラデを適用 ★）
# =========================
has_user_input = any(msg["role"]=="user" for msg in st.session_state["chat_history"])

if has_user_input:
    with st.container():
        # この“目印”が同じ stVerticalBlock 内にあると、上のCSSが当たる
        st.markdown('<div class="gen-scope"></div>', unsafe_allow_html=True)

        if st.button("AI見積もりくんで見積もりを生成する", key="gen_estimate"):
            with st.spinner("AIが見積もりを生成中…"):
                prompt = build_prompt_for_estimation(st.session_state["chat_history"])
                resp = openai_client.chat.completions.create(
                    model="gpt-4.1",
                    messages=[{"role":"system","content":"You MUST return only valid JSON."},
                              {"role":"user","content":prompt}],
                    response_format={"type":"json_object"},
                    temperature=0.2,
                    max_tokens=4000
                )
                raw = resp.choices[0].message.content or '{"items":[]}'
                items_json = robust_parse_items_json(raw)
                df = df_from_items_json(items_json)

                if df.empty:
                    st.warning("見積もりを出せませんでした。追加で要件を教えてください。")
                else:
                    meta = compute_totals(df)
                    st.session_state["items_json_raw"] = raw
                    st.session_state["items_json"] = items_json
                    st.session_state["df"] = df
                    st.session_state["meta"] = meta

                    # 入力欄の直上にヒント表示
                    hint_placeholder.caption(
                        "チャットをさらに続けて見積もり精度を上げることができます。\n"
                        "追加で要件を入力した後に再度このボタンを押すと、過去のチャット履歴＋新しい要件を反映して見積もりが更新されます。"
                    )

# =========================
# 表示 & ダウンロード
# =========================
if st.session_state["df"] is not None:
    st.markdown('<div class="preview-title">見積もり結果プレビュー</div>', unsafe_allow_html=True)
    st.dataframe(st.session_state["df"])
    st.write(f"**小計（税抜）:** {st.session_state['meta']['taxable']:,}円")
    st.write(f"**消費税:** {st.session_state['meta']['tax']:,}円")
    st.write(f"**合計:** {st.session_state['meta']['total']:,}円")

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        st.session_state["df"].to_excel(writer, index=False, sheet_name="見積もり")
    buf.seek(0)
    st.download_button("Excelでダウンロード", buf, "見積もり.xlsx",
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    tmpl = st.file_uploader("DD見積書テンプレートをアップロード（.xlsx）", type=["xlsx"])
    if tmpl is not None:
        out = export_with_template(tmpl.read(), st.session_state["df"])
        st.download_button("DD見積書テンプレで出力", out, "見積もり_DDテンプレ.xlsx")
